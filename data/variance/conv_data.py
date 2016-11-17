#! /usr/bin/env python

import sys
import os.path
import openpyxl

if(len(sys.argv) > 1):
	if(os.path.isfile(sys.argv[1])):
		wb = openpyxl.load_workbook(sys.argv[1], data_only=True)
		#wb = openpyxl.load_workbook('/Users/sdebruin/Dropbox/PowerBlade/PowerBlade_Variance.xlsx', data_only=True)
	else:
		print("File not found")
		exit()	
else:
	wb = openpyxl.load_workbook('/Users/sdebruin/repo/powerblade/data/variance/PowerBlade_Variance.xlsx', data_only=True)

sheet = wb.get_sheet_by_name('Sheet1')

col_name = 3

col_act = 9
col_norm = 16
col_intoJumper = 26
col_jumperInfo = 36
col_surge = 46

ci_offset = 4

print("plot ", end="")

testfile = open('outfile', 'a')

i = 1
for dev in range(1, sheet.max_row + 1):
	devname = sheet.cell(row = dev, column = col_name).value
	if(devname is not None and devname != "Load" and devname != 3):
		#print(devname)
		devname = devname.translate({ord(c): "" for c in "\""})
		devname = devname.translate({ord(c): "_" for c in "/\\ "})
		filename = 'abs' + str(devname) + '.dat'
		absfile = open(filename, 'w')
		#pctfile = open('pct' + str(devname) + '.dat', 'w')

		actPower = sheet.cell(row = dev, column = col_act).value

		# if(devname == 'Amplifier'):
		# 	print(actPower)
		# 	print(str(float(sheet.cell(row = dev, column = col_jumperInfo).value)))

		#i = 1
		# absfile.write(str(i) + ',Actual,' + str(sheet.cell(row = dev, column = col_act).value) + ',' + str(sheet.cell(row = dev, column = (col_act + 1)).value) + '\n')
		# i += 1
		try:
			absfile.write(str(i) + ',Outlet,' + str(float(sheet.cell(row = dev, column = col_norm).value)/actPower) + ',' + str(float(sheet.cell(row = dev, column = (col_norm + ci_offset)).value)/actPower) + '\n')
			i += 1
		except:
			pass
		try:
			absfile.write(str(i) + ',Into Jumper,' + str(float(sheet.cell(row = dev, column = col_intoJumper).value)/actPower) + ',' + str(float(sheet.cell(row = dev, column = (col_intoJumper + ci_offset)).value)/actPower) + '\n')
			i += 1
		except:
			pass
		try:
			absfile.write(str(i) + ',Jumper Info,' + str(float(sheet.cell(row = dev, column = col_jumperInfo).value)/actPower) + ',' + str(float(sheet.cell(row = dev, column = (col_jumperInfo + ci_offset)).value)/actPower) + '\n')
			i += 1
		except:
			pass
		try:
			absfile.write(str(i) + ',Into Surge,' + str(float(sheet.cell(row = dev, column = col_surge).value)/actPower) + ',' + str(float(sheet.cell(row = dev, column = (col_surge + ci_offset)).value)/actPower) + '\n')
			i += 1
		except:
			pass

		print('\'' + filename + '\' u 1:3:4 with errorbars ,\\\n\t', end="")
		print('\'' + filename + '\' u 1:3 with lines ,\\\n\t', end="")
		#testfile.write('\'' + filename + '\' u 1:3:4 with errorbars ,\\ \n\t')


